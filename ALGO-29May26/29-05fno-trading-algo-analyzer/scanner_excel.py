"""""
Websocket.xlsx helpers: quote columns, sheet validation, ITM rows, hourly signal CSV log.
Import-only module (no Excel open on import).
"""

import csv
import datetime
import os
import re
import time
from collections import Counter
from datetime import date
from typing import Dict, List, Optional, Set, Tuple

import pandas as pd

# --- Quote columns C:J ---
SHEET_QUOTE_HEADERS = [
    "LTP", "ATP", "Volume", "Total Sell Quantity", "Open", "Close", "High", "Low",
]
SHEET_OI_HEADERS = ["OI", "OI Change"]


def _num(val):
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def quote_row_from_dhan(response: dict) -> list:
    if not isinstance(response, dict):
        return [None] * 8
    ltp = response.get("LTP") or response.get("last_price") or response.get("ltp")
    ohlc = response.get("ohlc") if isinstance(response.get("ohlc"), dict) else {}
    return [
        _num(ltp),
        _num(
            response.get("avg_price")
            or response.get("average_price")
            or response.get("ATP")
        ),
        _num(response.get("volume")),
        _num(
            response.get("total_sell_quantity")
            or response.get("sell_quantity")
        ),
        _num(ohlc.get("open") or response.get("open")),
        _num(ohlc.get("close") or response.get("close")),
        _num(ohlc.get("high") or response.get("high")),
        _num(ohlc.get("low") or response.get("low")),
    ]


def merge_quote_rows(rest_row, chain_row):
    """Prefer REST/websocket full quote; fill LTP/volume from option chain if missing."""
    p = list(rest_row) if rest_row else [None] * 8
    s = list(chain_row) if chain_row else [None] * 8
    if len(p) < 8:
        p.extend([None] * (8 - len(p)))
    if len(s) < 8:
        s.extend([None] * (8 - len(s)))
    out = []
    for i in range(8):
        pv, sv = p[i], s[i]

        def _filled(v):
            if v is None or v == "":
                return False
            try:
                return float(v) != 0 or i == 0
            except (TypeError, ValueError):
                return True

        if _filled(pv):
            out.append(pv)
        elif _filled(sv):
            out.append(sv)
        else:
            out.append(pv if pv is not None else sv)
    return out


def quote_row_from_chain_option(option: dict) -> list:
    if not isinstance(option, dict):
        return [None] * 8
    return [
        _num(option.get("ltp") or option.get("last_price")),
        _num(option.get("avg_price")),
        _num(option.get("volume")),
        _num(option.get("total_sell_quantity")),
        _num(option.get("open")),
        _num(option.get("close")),
        _num(option.get("high")),
        _num(option.get("low")),
    ]


def build_quote_map_from_chain(chain) -> dict:
    out = {}
    if not chain or not isinstance(chain, dict):
        return out
    spot = chain.get("spot")
    for o in chain.get("options") or []:
        if not isinstance(o, dict) or not o.get("symbol"):
            continue
        sym = str(o["symbol"]).strip()
        row = quote_row_from_chain_option(o)
        out[sym] = row
        out[sym.upper()] = row
    if spot:
        idx_row = [_num(spot)] + [None] * 7
        und = str(chain.get("underlying") or "NIFTY").upper()
        keys = ("NIFTY", "NIFTY 50", "nifty", "nifty 50")
        if und == "SENSEX":
            keys = ("SENSEX", "sensex")
        elif und == "BANKEX":
            keys = ("BANKEX", "bankex")
        for k in keys:
            out[k] = idx_row
    return out


def build_oi_map_from_chain(chain) -> dict:
    """Per-symbol OI + day change from option chain API."""
    out = {}
    if not chain or not isinstance(chain, dict):
        return out
    for o in chain.get("options") or []:
        if not isinstance(o, dict) or not o.get("symbol"):
            continue
        sym = str(o["symbol"]).strip()
        oi = int(o.get("oi") or 0)
        prev = int(o.get("previous_oi") or 0)
        chg = int(o.get("oi_change") if o.get("oi_change") is not None else (oi - prev))
        row = [oi, chg]
        out[sym] = row
        out[sym.upper()] = row
    return out


def build_chain_maps(chain):
    return build_quote_map_from_chain(chain), build_oi_map_from_chain(chain)


def build_combined_chain(*chains):
    """Merge NIFTY + SENSEX (etc.) option legs for Excel sync / OI monitor."""
    options = []
    underlyings = {}
    cached_at = 0.0
    for chain in chains:
        if not chain or not isinstance(chain, dict):
            continue
        options.extend(chain.get("options") or [])
        und = chain.get("underlying")
        if und and chain.get("spot") is not None:
            underlyings[str(und)] = chain.get("spot")
        cached_at = max(cached_at, float(chain.get("_cached_at") or 0))
    if not options:
        return None
    return {
        "options": options,
        "underlyings": underlyings,
        "_cached_at": cached_at or time.time(),
    }


def ensure_sheet_quote_headers(sheet):
    try:
        existing = sheet.range("C1").value
        if existing is None or str(existing).strip() == "":
            sheet.range("C1:J1").value = [SHEET_QUOTE_HEADERS]
        oi_h = sheet.range("K1").value
        if oi_h is None or str(oi_h).strip() == "":
            sheet.range("K1:L1").value = [SHEET_OI_HEADERS]
    except Exception:
        pass


# --- Sheet validator ---
_MONTHS = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12,
}
_INDEX_NSE = {
    "NIFTY", "NIFTY 50", "BANKNIFTY", "NIFTY BANK", "FINNIFTY",
    "NIFTY FIN SERVICE", "MIDCPNIFTY", "NIFTY MID SELECT",
}
_INDEX_BSE = {"SENSEX", "BANKEX"}
_UNDERLYING_IDS = {
    "NIFTY": (13, "IDX_I"), "BANKNIFTY": (25, "IDX_I"),
    "FINNIFTY": (27, "IDX_I"), "MIDCPNIFTY": (442, "IDX_I"),
    "SENSEX": (51, "IDX_I"), "BANKEX": (69, "IDX_I"),
}
_UNDERLYING_ALIASES = {
    "NIFTY 50": "NIFTY", "NIFTY BANK": "BANKNIFTY",
    "NIFTY FIN SERVICE": "FINNIFTY", "NIFTY MID SELECT": "MIDCPNIFTY",
}
FNO_RE = re.compile(
    r"^(.+?)\s+(\d{1,2})\s+(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)\s+"
    r"(\d+(?:\.\d+)?)\s+(CALL|PUT|CE|PE)$",
    re.IGNORECASE,
)
_BSE_FNO_RE = re.compile(
    r"^(\d{4}\s+)?(SENSEX|BANKEX)\s+(\d{1,2})\s+"
    r"(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)\s+"
    r"(\d+(?:\.\d+)?)(?:\s+(CALL|PUT|CE|PE))?$",
    re.IGNORECASE,
)
INSTRUMENT_EXCHANGE = {
    "NSE": "NSE", "BSE": "BSE", "NFO": "NSE", "BFO": "BSE", "MCX": "MCX",
    "CUR": "NSE", "BSE_IDX": "BSE", "NSE_IDX": "NSE",
}


def _normalize_exchange(symbol: str, exchange: str) -> str:
    sym = str(symbol).upper().strip()
    exch = str(exchange or "NSE").strip().upper()
    if sym in _INDEX_NSE:
        return "NSE_IDX"
    if sym in _INDEX_BSE:
        return "BSE_IDX"
    return exch


def _parse_fno_symbol(symbol: str) -> Optional[dict]:
    m = FNO_RE.match(" ".join(str(symbol).upper().split()))
    if not m:
        return None
    und, day, month, strike, otype = m.groups()
    und_key = _UNDERLYING_ALIASES.get(und.strip().upper(), und.strip().upper())
    return {
        "underlying": und_key,
        "token": f"{int(day)} {month.upper()}",
    }


def _expiry_to_tokens(exp) -> Set[str]:
    tokens = set()
    try:
        d = pd.to_datetime(exp).date()
        tokens.add(d.isoformat())
        for k, v in _MONTHS.items():
            if v == d.month:
                tokens.add(f"{d.day} {k}")
                break
    except Exception:
        pass
    return tokens


def fetch_active_expiries(tsl, underlyings=None) -> Dict[str, Set[str]]:
    if underlyings is None:
        underlyings = tuple(_UNDERLYING_IDS.keys())
    out: Dict[str, Set[str]] = {}
    for u in underlyings:
        if u not in _UNDERLYING_IDS:
            continue
        sid, seg = _UNDERLYING_IDS[u]
        try:
            resp = tsl.Dhan.expiry_list(sid, seg)
            block = resp.get("data") if isinstance(resp, dict) else None
            if isinstance(block, dict) and isinstance(block.get("data"), list):
                block = block["data"]
            if isinstance(block, list) and block:
                allowed = set()
                for exp in block[:4]:
                    allowed |= _expiry_to_tokens(exp)
                out[u] = allowed
        except Exception:
            continue
    return out


def validate_row(symbol, exchange, instrument_df, active_expiries, chain_symbols=None):
    sym = str(symbol).strip()
    if not sym or sym.lower() == "nan":
        return False, "empty symbol"
    matched = instrument_df[
        (
            (instrument_df["SEM_TRADING_SYMBOL"] == sym)
            | (instrument_df["SEM_CUSTOM_SYMBOL"] == sym)
        )
        & (instrument_df["SEM_EXM_EXCH_ID"] == INSTRUMENT_EXCHANGE.get(
            _normalize_exchange(sym, exchange), "NSE"
        ))
    ]
    if _normalize_exchange(sym, exchange) in ("NSE", "BSE") and "SEM_SERIES" in matched.columns:
        eq = matched[matched["SEM_SERIES"] == "EQ"]
        if not eq.empty:
            matched = eq
    row = matched.iloc[-1] if not matched.empty else None
    parsed = _parse_fno_symbol(sym)
    if row is None and parsed is None:
        return False, "no instrument match"
    is_index = sym.upper() in _INDEX_NSE or sym.upper() in _INDEX_BSE
    if row is not None and not is_index and "SEM_EXPIRY_DATE" in row.index:
        exp_raw = row.get("SEM_EXPIRY_DATE")
        if pd.notna(exp_raw):
            try:
                exp_d = pd.to_datetime(exp_raw).date()
                if 2000 <= exp_d.year and exp_d < date.today():
                    return False, f"expired ({exp_d})"
            except Exception:
                pass
    if parsed is not None:
        allowed = active_expiries.get(parsed["underlying"], set())
        if allowed and parsed["token"] not in allowed:
            return False, f"stale expiry ({parsed['token']})"
    if row is None:
        return False, "no instrument match"
    return True, "ok"


def clean_rows(rows, instrument_df, active_expiries, chain_symbols=None):
    keep, deleted = [], []
    for sym, exch in rows:
        ok, reason = validate_row(sym, exch, instrument_df, active_expiries, chain_symbols)
        if ok:
            keep.append((sym, exch))
        else:
            deleted.append((sym, reason))
    return keep, deleted


# --- Strike window: 3 CE + 3 PE from ATM (Websocket.xlsx) ---
STRIKES_EACH_SIDE = 3
INDEX_SYMBOLS = set(_INDEX_NSE) | set(_INDEX_BSE) | {"NIFTY", "NIFTY 50"}
_last_strike_syms_by_und = {}
_last_strike_refresh_by_und = {}
_strike_log_last = 0.0


def is_fno_symbol(symbol: str) -> bool:
    s = str(symbol or "").strip()
    if not s:
        return False
    if FNO_RE.match(s):
        return True
    return bool(_BSE_FNO_RE.match(s))


def underlying_step(chain) -> int:
    und = str((chain or {}).get("underlying") or "NIFTY").upper()
    if und in ("SENSEX", "BANKEX"):
        return 100
    if und in ("BANKNIFTY", "NIFTY BANK"):
        return 100
    return 50


def fno_exchange_for_symbol(sym: str) -> str:
    s = str(sym or "").upper()
    if "SENSEX" in s or "BANKEX" in s:
        return "BFO"
    return "NFO"


def select_options_near_atm(chain, n_ce=STRIKES_EACH_SIDE, n_pe=STRIKES_EACH_SIDE, step=None):
    """ATM CE + lower strikes; ATM PE + higher strikes (3 each by default)."""
    if not chain or not isinstance(chain, dict):
        return []
    atm = int(chain.get("atm") or 0)
    if step is None:
        step = underlying_step(chain)
    if not atm:
        spot = float(chain.get("spot") or 0)
        atm = round(spot / step) * step if spot else 0
    if not atm:
        return []
    by_key = {}
    for o in chain.get("options") or []:
        if not isinstance(o, dict):
            continue
        st = int(o.get("strike") or 0)
        ot = o.get("option_type")
        if ot not in ("CE", "PE"):
            continue
        key = (st, ot)
        if key not in by_key or int(o.get("volume") or 0) > int(by_key[key].get("volume") or 0):
            by_key[key] = o
    selected = []
    for i in range(n_ce):
        o = by_key.get((atm - i * step, "CE"))
        if o:
            selected.append(o)
    for i in range(n_pe):
        o = by_key.get((atm + i * step, "PE"))
        if o:
            selected.append(o)
    return selected


def filter_chain_strike_window(chain, n_ce=STRIKES_EACH_SIDE, n_pe=STRIKES_EACH_SIDE, step=None):
    if not chain or not isinstance(chain, dict):
        return chain
    if step is None:
        step = underlying_step(chain)
    out = dict(chain)
    out["options"] = select_options_near_atm(chain, n_ce, n_pe, step)
    out["strike_window"] = {"ce": n_ce, "pe": n_pe, "step": step}
    return out


def strike_window_symbols(chain, n_ce=STRIKES_EACH_SIDE, n_pe=STRIKES_EACH_SIDE, step=None):
    if step is None:
        step = underlying_step(chain)
    return {
        str(o.get("symbol", "")).strip()
        for o in select_options_near_atm(chain, n_ce, n_pe, step)
        if o.get("symbol")
    }


def refresh_strike_window_on_websocket(
    sheet, chain, *, refresh_sec=90, prune_extra_fno=True,
):
    """Keep 3 CE + 3 PE from ATM per underlying; BFO for SENSEX/BANKEX."""
    global _last_strike_syms_by_und, _last_strike_refresh_by_und, _strike_log_last
    if sheet is None or not chain:
        return 0
    underlying = str(chain.get("underlying") or "NIFTY").upper()
    now = time.time()
    last_refresh = _last_strike_refresh_by_und.get(underlying, 0.0)
    if now - last_refresh < refresh_sec:
        return 0
    _last_strike_refresh_by_und[underlying] = now
    step = underlying_step(chain)
    window_opts = select_options_near_atm(chain, step=step)
    new_syms = {str(o.get("symbol", "")).strip() for o in window_opts if o.get("symbol")}
    new_upper = {s.upper() for s in new_syms}
    prev_auto = _last_strike_syms_by_und.get(underlying, set())
    prev_upper = {s.upper() for s in prev_auto}
    if not new_syms and not prev_auto:
        return 0
    try:
        last_row = sheet.range("A1").end("down").row
        df = sheet.range(f"A1:B{max(last_row, 2)}").options(
            pd.DataFrame, header=1, index=False
        ).value
    except Exception as exc:
        print(f"[!] Strike window sheet read failed: {exc}", flush=True)
        return 0
    name_col = "Script Name" if "Script Name" in df.columns else df.columns[0]
    rows_to_delete, existing_upper = [], set()
    if df is not None and not df.empty:
        for i in range(len(df)):
            sym = str(df[name_col].iloc[i]).strip()
            if not sym or sym.lower() == "nan":
                continue
            existing_upper.add(sym.upper())
            in_window = sym in new_syms or sym.upper() in new_upper
            if sym.upper() in INDEX_SYMBOLS:
                continue
            was_auto = sym.upper() in prev_upper
            if prune_extra_fno and was_auto and is_fno_symbol(sym) and not in_window:
                rows_to_delete.append(i + 2)
    for row_num in sorted(rows_to_delete, reverse=True):
        try:
            sheet.api.Rows(row_num).Delete()
        except Exception:
            pass
    added = 0
    append_row = sheet.range("A1").end("down").row + 1
    if append_row < 2:
        append_row = 2
    for o in window_opts:
        sym = str(o.get("symbol", "")).strip()
        if not sym or sym.upper() in existing_upper:
            continue
        try:
            sheet.range(f"A{append_row}").value = sym
            sheet.range(f"B{append_row}").value = fno_exchange_for_symbol(sym)
            existing_upper.add(sym.upper())
            append_row += 1
            added += 1
        except Exception:
            break
    _last_strike_syms_by_und[underlying] = new_syms
    if now - _strike_log_last >= 60 and new_syms:
        _strike_log_last = now
        ce_n = sum(1 for o in window_opts if o.get("option_type") == "CE")
        pe_n = sum(1 for o in window_opts if o.get("option_type") == "PE")
        removed = len(rows_to_delete)
        print(
            f"[i] Websocket {underlying} strike window: {ce_n} CE + {pe_n} PE @ step {step} "
            f"(removed {removed} stale auto row(s), +{added} new).",
            flush=True,
        )
    return added


def refresh_itm_rows_on_websocket(
    sheet, chain, spot, previous_close, today_open, *, refresh_sec=90,
):
    """Backward-compatible alias — uses 3 CE + 3 PE window, not deep ITM list."""
    return refresh_strike_window_on_websocket(sheet, chain, refresh_sec=refresh_sec)


# --- Hourly signal CSV log (IST session 09:30–15:15) ---
LOG_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Dependencies", "hourly_logs")
LOG_SESSION_START = datetime.time(9, 30)
LOG_SESSION_END = datetime.time(15, 15)
LOG_SESSION_HARD_STOP = datetime.time(15, 30)
_HOURLY_FIELDS = [
    "timestamp", "event", "symbol", "option_type", "strike", "ltp", "score",
    "required_score", "rsi", "trend", "pcr", "nifty_spot", "gap_mode",
    "scanner_state", "reject_reason", "notes",
]
_TRACKED_EVENTS = frozenset({"SIGNAL_VERIFIED", "REJECT"})


class HourlySignalLogger:
    def __init__(self):
        self._hour_key = None
        self._csv_path = None
        self._hour_has_activity = False
        self._session_closed = False

    @staticmethod
    def current_hour_key():
        return time.strftime("%Y-%m-%d_%H00")

    @staticmethod
    def csv_path_for_hour(hour_key=None):
        hk = hour_key or HourlySignalLogger.current_hour_key()
        return os.path.join(LOG_DIR, f"signal_hourly_{hk}.csv")

    @staticmethod
    def summary_path_for_hour(hour_key):
        return os.path.join(LOG_DIR, f"hourly_summary_{hour_key}.txt")

    def log_dir(self):
        os.makedirs(LOG_DIR, exist_ok=True)
        return LOG_DIR

    def _summarize_csv(self, csv_path):
        if not csv_path or not os.path.isfile(csv_path):
            return Counter(), []
        counts = Counter()
        rows = []
        with open(csv_path, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                counts[row.get("event") or "UNKNOWN"] += 1
                rows.append(row)
        return counts, rows

    def _emit_hour_summary(self, hour_key, csv_path):
        counts, rows = self._summarize_csv(csv_path)
        os.makedirs(LOG_DIR, exist_ok=True)
        summary_path = self.summary_path_for_hour(hour_key)
        lines = [
            f"Hour bucket: {hour_key} (IST wall-clock hour)",
            f"Detail CSV: {csv_path}",
            f"Total events: {sum(counts.values())}",
        ]
        for ev, n in sorted(counts.items()):
            lines.append(f"  {ev}: {n}")
        signals = [r for r in rows if r.get("event") == "SIGNAL_VERIFIED"]
        rejects = [r for r in rows if r.get("event") == "REJECT"]
        if signals:
            lines.append("Signals:")
            for r in signals:
                lines.append(
                    f"  {r.get('timestamp')} {r.get('symbol')} score={r.get('score')} "
                    f"RSI={r.get('rsi')} trend={r.get('trend')}"
                )
        if rejects:
            lines.append("Top rejects (score>=85):")
            for r in rejects[:20]:
                lines.append(
                    f"  {r.get('timestamp')} {r.get('symbol')} score={r.get('score')} "
                    f"— {r.get('reject_reason')}"
                )
            if len(rejects) > 20:
                lines.append(f"  ... +{len(rejects) - 20} more in CSV")
        text = "\n".join(lines) + "\n"
        with open(summary_path, "w", encoding="utf-8") as f:
            f.write(text)
        print("\n" + "=" * 60, flush=True)
        print(f"[hourly] Past hour closed — {hour_key}", flush=True)
        for line in lines:
            print(f"[hourly] {line}", flush=True)
        print(f"[hourly] Summary saved: {summary_path}", flush=True)
        print("=" * 60 + "\n", flush=True)

    def print_startup_info(self):
        os.makedirs(LOG_DIR, exist_ok=True)
        hk = self.current_hour_key()
        print("[i] Hypercare hourly report folder:", flush=True)
        print(f"    {LOG_DIR}", flush=True)
        print(
            "    One CSV per hour (09:30–15:15 IST): signal_hourly_YYYY-MM-DD_HH00.csv",
            flush=True,
        )
        print(
            "    Each file has HOUR_START; NO_UPDATE if no signals that hour.",
            flush=True,
        )
        if self._logging_active():
            self._open_hour_file(hk)

    @staticmethod
    def _logging_active(now=None):
        now = now or datetime.datetime.now()
        if now.weekday() >= 5:
            return False
        t = now.time()
        return LOG_SESSION_START <= t <= LOG_SESSION_HARD_STOP

    def _should_log_hour(self, hour_key, now=None):
        now = now or datetime.datetime.now()
        if not self._logging_active(now):
            return False
        try:
            hour = int(hour_key.split("_")[1][:2])
        except (IndexError, ValueError):
            return False
        if hour < 9 or hour > 15:
            return False
        if hour == 9 and now.time() < LOG_SESSION_START:
            return False
        if hour > 15:
            return False
        return True

    def _write_row(self, event, **kwargs):
        if not self._csv_path:
            return
        row = {"timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "event": event}
        row.update({k: kwargs.get(k, "") for k in _HOURLY_FIELDS if k not in row})
        with open(self._csv_path, "a", newline="", encoding="utf-8") as f:
            csv.DictWriter(f, fieldnames=_HOURLY_FIELDS).writerow(row)

    def _open_hour_file(self, hour_key):
        os.makedirs(LOG_DIR, exist_ok=True)
        self._csv_path = self.csv_path_for_hour(hour_key)
        new_file = not os.path.isfile(self._csv_path)
        if new_file:
            with open(self._csv_path, "w", newline="", encoding="utf-8") as f:
                csv.DictWriter(f, fieldnames=_HOURLY_FIELDS).writeheader()
            self._write_row(
                "HOUR_START",
                notes="Hourly hypercare log active; waiting for signals/rejects.",
            )
            status_path = os.path.join(LOG_DIR, f"hourly_status_{hour_key}.txt")
            with open(status_path, "w", encoding="utf-8") as f:
                f.write(
                    f"Hour {hour_key} started at "
                    f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} IST\n"
                    "Status: monitoring active\n"
                )
            print(f"[hourly] New file: {self._csv_path}", flush=True)
        self._hour_has_activity = False

    def _close_hour(self, hour_key, csv_path):
        prev_path = self._csv_path
        self._csv_path = csv_path
        counts, _ = self._summarize_csv(csv_path)
        activity = sum(counts.get(e, 0) for e in _TRACKED_EVENTS)
        if activity == 0:
            self._write_row(
                "NO_UPDATE",
                notes="No signal verified and no score-85 reject this hour.",
            )
            status_path = os.path.join(LOG_DIR, f"hourly_status_{hour_key}.txt")
            with open(status_path, "a", encoding="utf-8") as f:
                f.write(
                    f"\nClosed {datetime.datetime.now().strftime('%H:%M:%S')}: "
                    "NO_UPDATE (no signals/rejects logged).\n"
                )
        self._emit_hour_summary(hour_key, csv_path)
        self._csv_path = prev_path

    def tick(self, scanner_state="SCAN", **kwargs):
        """Ensure current-hour CSV exists during session; close hours after 15:15."""
        now = datetime.datetime.now()
        if self._session_closed and now.time() > LOG_SESSION_END:
            return
        if not self._logging_active(now):
            return
        hour_key = self.current_hour_key()
        if now.time() > LOG_SESSION_END:
            if self._hour_key and self._csv_path:
                self._close_hour(self._hour_key, self._csv_path)
                self._hour_key = None
                self._csv_path = None
            self._session_closed = True
            return
        if not self._should_log_hour(hour_key, now):
            return
        if self._hour_key is not None and self._hour_key != hour_key and self._csv_path:
            self._close_hour(self._hour_key, self._csv_path)
        if self._hour_key != hour_key:
            self._hour_key = hour_key
            self._open_hour_file(hour_key)

    def log(self, event, **kwargs):
        now = datetime.datetime.now()
        if not self._logging_active(now):
            return
        hour_key = self.current_hour_key()
        if not self._should_log_hour(hour_key, now):
            return
        if self._hour_key is not None and self._hour_key != hour_key and self._csv_path:
            self._close_hour(self._hour_key, self._csv_path)
        if self._hour_key != hour_key:
            self._hour_key = hour_key
            self._open_hour_file(hour_key)
        if event in _TRACKED_EVENTS:
            self._hour_has_activity = True
        self._write_row(event, **kwargs)


hourly_log = HourlySignalLogger()


# --- OI / Volume alert log (separate Excel, openpyxl — not via websocket) ---
OI_VOL_XLSX = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "OI_Volume_Alerts.xlsx"
)
_LOG_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Dependencies")
OI_VOL_CSV = os.path.join(_LOG_DIR, "oi_volume_alerts.csv")


def _oi_vol_today_csv():
    day = datetime.datetime.now().strftime("%Y-%m-%d")
    return os.path.join(_LOG_DIR, f"oi_volume_alerts_{day}.csv")


def _oi_vol_csv_paths():
    """Today's file first (always used); legacy file mirrored when writable."""
    paths = [_oi_vol_today_csv()]
    if OI_VOL_CSV not in paths:
        paths.append(OI_VOL_CSV)
    return paths

OI_VOL_HEADERS = [
    "timestamp",
    "alert_type",
    "symbol",
    "oi",
    "prev_oi",
    "oi_change_pct",
    "oi_delta",
    "volume_1m",
    "volume_ema_20",
    "cross_direction",
    "message",
]


class OiVolumeAlertLogger:
    """Append [oi] / [vol] console alerts for later review (Excel file)."""

    def __init__(self):
        self._last_save_err = 0.0
        self._locked_warned = False
        self._legacy_csv_warned = False
        self._last_snapshot_ts = 0.0
        self._last_poll_oi = {}
        self._last_poll_log = {}

    def path(self):
        return OI_VOL_XLSX

    def csv_path(self):
        return _oi_vol_today_csv()

    def _ensure_csv_header(self, path):
        os.makedirs(_LOG_DIR, exist_ok=True)
        if os.path.isfile(path) and os.path.getsize(path) > 0:
            return
        with open(path, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow(OI_VOL_HEADERS)

    def _write_csv_row(self, path, row):
        new_file = not os.path.isfile(path) or os.path.getsize(path) == 0
        with open(path, "a", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            if new_file:
                w.writerow(OI_VOL_HEADERS)
            w.writerow(row)
            f.flush()
            try:
                os.fsync(f.fileno())
            except OSError:
                pass

    def _append_csv(self, row):
        os.makedirs(_LOG_DIR, exist_ok=True)
        today = _oi_vol_today_csv()
        self._ensure_csv_header(today)
        last_err = None
        wrote_today = False
        for path in _oi_vol_csv_paths():
            try:
                self._write_csv_row(path, row)
                if path == today:
                    wrote_today = True
                elif path == OI_VOL_CSV and not getattr(self, "_legacy_csv_warned", False):
                    self._legacy_csv_warned = True
            except PermissionError as exc:
                last_err = exc
                if path == today:
                    break
                continue
        if not wrote_today:
            if last_err:
                raise last_err
            raise PermissionError(f"Cannot write OI/Volume log: {today}")
        return today

    def _backup_corrupt_xlsx(self, exc):
        stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = f"{OI_VOL_XLSX}.corrupt_{stamp}"
        try:
            os.replace(OI_VOL_XLSX, backup)
        except OSError:
            backup = None
        now = time.time()
        if now - self._last_save_err > 120:
            self._last_save_err = now
            suffix = f" Backup: {backup}" if backup else ""
            print(
                f"[!] OI_Volume_Alerts.xlsx was corrupt — recreating ({exc}).{suffix}",
                flush=True,
            )

    def _seed_xlsx_from_csv(self, ws):
        for path in _oi_vol_csv_paths():
            if not os.path.isfile(path) or os.path.getsize(path) == 0:
                continue
            try:
                with open(path, newline="", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    next(reader, None)
                    for row in reader:
                        ws.append(row)
                return True
            except OSError:
                continue
        return False

    def _load_or_create_xlsx(self):
        from openpyxl import Workbook, load_workbook
        from zipfile import BadZipFile

        if os.path.isfile(OI_VOL_XLSX):
            try:
                wb = load_workbook(OI_VOL_XLSX)
                ws = wb["Alerts"] if "Alerts" in wb.sheetnames else wb.active
                return wb, ws
            except (BadZipFile, KeyError, OSError) as exc:
                self._backup_corrupt_xlsx(exc)

        wb = Workbook()
        ws = wb.active
        ws.title = "Alerts"
        ws.append(OI_VOL_HEADERS)
        self._seed_xlsx_from_csv(ws)
        return wb, ws

    def _append_xlsx(self, row):
        wb, ws = self._load_or_create_xlsx()
        ws.append(row)
        tmp_path = f"{OI_VOL_XLSX}.tmp"
        try:
            wb.save(tmp_path)
            os.replace(tmp_path, OI_VOL_XLSX)
        finally:
            if os.path.isfile(tmp_path):
                try:
                    os.remove(tmp_path)
                except OSError:
                    pass

    def _xlsx_is_locked(self):
        if not os.path.isfile(OI_VOL_XLSX):
            return False
        try:
            with open(OI_VOL_XLSX, "a+b"):
                pass
            return False
        except PermissionError:
            return True
        except OSError:
            return True

    def append(self, alert_type, symbol, message, **fields):
        row = [
            datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            alert_type,
            symbol,
            fields.get("oi", ""),
            fields.get("prev_oi", ""),
            fields.get("oi_change_pct", ""),
            fields.get("oi_delta", ""),
            fields.get("volume_1m", ""),
            fields.get("volume_ema_20", ""),
            fields.get("cross_direction", ""),
            message,
        ]
        try:
            self._append_csv(row)
        except PermissionError:
            if not self._locked_warned:
                self._locked_warned = True
                print(
                    f"[!] OI/Volume log locked — close in Excel: {_oi_vol_today_csv()} "
                    f"and {OI_VOL_XLSX}",
                    flush=True,
                )
            return
        try:
            self._append_xlsx(row)
        except PermissionError:
            if not self._locked_warned:
                self._locked_warned = True
                print(
                    "[!] OI_Volume_Alerts.xlsx is OPEN in Excel — rows still go to CSV.",
                    flush=True,
                )
        except Exception as exc:
            now = time.time()
            if now - self._last_save_err > 120:
                self._last_save_err = now
                print(f"[!] OI_Volume_Alerts.xlsx save failed: {exc}", flush=True)

    def print_startup_info(self):
        today_csv = _oi_vol_today_csv()
        self._ensure_csv_header(today_csv)
        print("[i] OI / Volume alert log (separate from Websocket.xlsx):", flush=True)
        print(f"    Excel: {OI_VOL_XLSX}", flush=True)
        print(f"    CSV (today — open this): {today_csv}", flush=True)
        print(f"    CSV (legacy mirror): {OI_VOL_CSV}", flush=True)
        if self._xlsx_is_locked():
            print(
                "[!] OI_Volume_Alerts.xlsx is locked — close it in Excel or use the CSV path above.",
                flush=True,
            )
        elif not os.path.isfile(OI_VOL_XLSX):
            print("[i] OI_Volume_Alerts.xlsx will be created on first [oi]/[vol] alert.", flush=True)
        print(
            "    Rows append on [oi]/[vol] alerts + Vol every 5 min while scanner runs.",
            flush=True,
        )


oi_volume_log = OiVolumeAlertLogger()

OI_VOL_SNAPSHOT_SEC = 300
OI_VOL_POLL_PCT = 5.0


def format_signed_int(value) -> str:
    """e.g. +392,580 or -120,000 for quick UP/DOWN scan in CSV."""
    try:
        n = int(value)
    except (TypeError, ValueError):
        return str(value)
    return f"{n:+,}"


def format_oi_poll_message(oi, prev_oi, pct_abs=None):
    """OI poll line with explicit +/- on % and delta."""
    oi = int(oi or 0)
    prev_oi = int(prev_oi or 0)
    delta = oi - prev_oi
    if pct_abs is None:
        pct_signed = (delta / prev_oi * 100.0) if prev_oi > 0 else 0.0
    else:
        pct_signed = float(pct_abs) if delta >= 0 else -float(pct_abs)
    return (
        f"OI poll change {pct_signed:+.1f}% vs last scan "
        f"(OI chg {format_signed_int(delta)} | oi={oi:,}, was={prev_oi:,})"
    )


def format_vol_snapshot_message(vol, oi, prev_oi):
    vol = int(vol or 0)
    oi = int(oi or 0)
    prev_oi = int(prev_oi or 0)
    if prev_oi > 0:
        delta = oi - prev_oi
        pct_signed = delta / prev_oi * 100.0
        return (
            f"periodic record vol={vol:,} oi={oi:,} "
            f"(OI chg {format_signed_int(delta)} {pct_signed:+.1f}%)"
        )
    return f"periodic record vol={vol:,} oi={oi:,}"


def run_oi_volume_monitor(chain):
    """
    Run OI/vol checks on every option in chain (not filtered by Websocket symbol names).
    Does not change Option_strategy_core signal rules.
    """
    import Option_strategy_core as core

    if not chain or not isinstance(chain, dict):
        return 0
    options = chain.get("options") or []
    if not options:
        return 0
    n_alert = 0
    now = time.time()

    for o in options:
        if not isinstance(o, dict) or not o.get("symbol"):
            continue
        sym = str(o["symbol"]).strip()
        vol = int(o.get("volume") or 0)
        oi = int(o.get("oi") or 0)
        prev_oi = int(o.get("previous_oi") or 0)
        if prev_oi <= 0 and oi > 0:
            prev_oi = max(oi - int(o.get("oi_change") or 0), 0)

        if core.check_volume_ema_cross(sym, vol):
            n_alert += 1
            try:
                from live_alert_logger import write_live_alert
                write_live_alert("VOL", f"{sym} crossed Volume EMA.")
            except Exception:
                pass
            play_alert_sound("vol")
        if core.check_oi_change_alert(o):
            n_alert += 1
            try:
                from live_alert_logger import write_live_alert
                write_live_alert("OI", f"{sym} has significant OI change.")
            except Exception:
                pass
            play_alert_sound("vol")

        last_oi = oi_volume_log._last_poll_oi.get(sym)
        oi_volume_log._last_poll_oi[sym] = oi
        if last_oi and last_oi > 0 and oi > 0:
            delta_poll = oi - last_oi
            pct_poll = abs(delta_poll) / last_oi * 100.0
            if pct_poll >= OI_VOL_POLL_PCT:
                last_t = oi_volume_log._last_poll_log
                if now - last_t.get(sym, 0) >= 60:
                    last_t[sym] = now
                    oi_volume_log._last_poll_log = last_t
                    pct_signed = (delta_poll / last_oi * 100.0) if last_oi else 0.0
                    msg = format_oi_poll_message(oi, last_oi, pct_abs=pct_signed)
                    try:
                        from live_alert_logger import write_live_alert
                        write_live_alert("OI", f"{sym}: {msg}")
                    except Exception:
                        pass
                    direction = "ABOVE" if delta_poll >= 0 else "BELOW"
                    oi_volume_log.append(
                        "Reject",
                        sym,
                        msg,
                        oi=oi,
                        prev_oi=last_oi,
                        oi_change_pct=round(pct_signed, 1),
                        oi_delta=delta_poll,
                        cross_direction=direction,
                    )
                    n_alert += 1

    if now - oi_volume_log._last_snapshot_ts >= OI_VOL_SNAPSHOT_SEC:
        oi_volume_log._last_snapshot_ts = now
        for o in options:
            if not isinstance(o, dict) or not o.get("symbol"):
                continue
            sym = str(o["symbol"]).strip()
            oi = int(o.get("oi") or 0)
            prev_oi = int(o.get("previous_oi") or 0)
            vol = int(o.get("volume") or 0)
            delta_snap = oi - prev_oi if prev_oi > 0 else 0
            pct_signed = round(delta_snap / prev_oi * 100.0, 1) if prev_oi > 0 else ""
            oi_volume_log.append(
                "Vol",
                sym,
                format_vol_snapshot_message(vol, oi, prev_oi),
                oi=oi,
                prev_oi=prev_oi,
                oi_change_pct=pct_signed,
                oi_delta=delta_snap,
                volume_1m=vol,
            )
        print(
            f"[i] OI/Vol log: {len(options)} Vol row(s) written "
            f"({oi_volume_log.csv_path()})",
            flush=True,
        )

    return n_alert


# --- Windows sound on BUY / SELL (no change to signal logic) ---
def play_alert_sound(alert_type="buy"):
    """Disabled by user via prompt."""
    return

def print_alert_sound_help():
    print("[i] Alert sounds: Disabled.", flush=True)
